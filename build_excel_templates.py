import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create excels directory
output_dir = "/home/ramji/Desktop/GRE/gre/excels"
os.makedirs(output_dir, exist_ok=True)

headers = [
    "Category", "Level", "Question Type", "Question Text", "Passage",
    "Option A", "Option B", "Option C", "Option D", "Option E", "Option F",
    "Correct Answer", "Explanation", "Question Images", "Answer Images"
]

def create_excel(filename, sheet_title, data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Write Header
    ws.append(headers)

    # Write Data Rows
    for row in data_rows:
        while len(row) < len(headers):
            row.append("")
        ws.append(row)

    # Style Header & Cells
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28
    wb.save(os.path.join(output_dir, filename))

# 1. AWA Questions Excel (10 rows)
awa_rows = [
    ["AWA Issue", "Medium", "AWA", "Educational institutions should make all courses optional for college students. Discuss the extent to which you agree or disagree.", "", "", "", "", "", "", "", "N/A", "Analyze student autonomy versus curriculum structure."],
    ["AWA Argument", "Hard", "AWA", "The following appeared in a memo from a company vice president: 'Our sales rose by 20% after implementing remote work, so all employees should work remotely.' Discuss the logical flaws.", "", "", "", "", "", "", "", "N/A", "Identify flaws in causal link between remote work and sales increase."],
    ["AWA Issue", "Easy", "AWA", "Governments should place higher priority on funding basic scientific research than on applied scientific research. Discuss your views.", "", "", "", "", "", "", "", "N/A", "Compare long-term foundational discovery vs immediate practical application."],
    ["AWA Argument", "Medium", "AWA", "The following appeared in a health journal: 'People who drink green tea daily report lower stress. Therefore, green tea cures anxiety.' Evaluate the argument.", "", "", "", "", "", "", "", "N/A", "Address correlation vs causation fallacy."],
    ["AWA Issue", "Hard", "AWA", "True greatness in any field can only be recognized long after a person's era has passed. Discuss the extent to which you agree.", "", "", "", "", "", "", "", "N/A", "Examine historical consensus vs contemporary recognition."],
    ["AWA Argument", "Easy", "AWA", "A city official claims: 'Installing traffic cameras reduced accidents by 15% on Main Street, so we should install them on all streets.' Evaluate the recommendation.", "", "", "", "", "", "", "", "N/A", "Consider sample bias and specific street characteristics."],
    ["AWA Issue", "Medium", "AWA", "In order to be an effective leader, a public official must maintain strict moral standards in private life. Discuss your perspective.", "", "", "", "", "", "", "", "N/A", "Explore private morality vs public competence."],
    ["AWA Argument", "Hard", "AWA", "The president of a gym chain asserts: 'Adding sauna rooms in Branch A doubled memberships, so adding saunas in all branches will double total revenue.' Evaluate.", "", "", "", "", "", "", "", "N/A", "Identify false generalization across different locations."],
    ["AWA Issue", "Easy", "AWA", "Formal education should primarily focus on preparing students for specific careers rather than providing general knowledge. Discuss.", "", "", "", "", "", "", "", "N/A", "Debate vocational specialization vs liberal arts breadth."],
    ["AWA Argument", "Medium", "AWA", "A housing board report states: 'Rent prices in Westville dropped 10% after building 500 apartments. Building 500 more will reduce prices by another 10%.' Evaluate.", "", "", "", "", "", "", "", "N/A", "Examine diminishing returns and supply-demand equilibrium."]
]
create_excel("sample_awa_questions.xlsx", "AWA Questions", awa_rows)

# 2. Quant Questions Excel (10 rows)
quant_rows = [
    ["GRE Quant QC", "Easy", "QUANTITATIVE_COMPARISON", "Quantity A: 2^5\nQuantity B: 5^2", "", "Quantity A is greater.", "Quantity B is greater.", "The two quantities are equal.", "The relationship cannot be determined from the information given.", "", "", "A", "2^5 = 32 and 5^2 = 25. 32 > 25, so Quantity A is greater."],
    ["GRE Quant QC", "Medium", "QUANTITATIVE_COMPARISON", "x > 0.\nQuantity A: x^2 + 1\nQuantity B: (x + 1)^2", "", "Quantity A is greater.", "Quantity B is greater.", "The two quantities are equal.", "The relationship cannot be determined from the information given.", "", "", "B", "(x+1)^2 = x^2 + 2x + 1. Since x > 0, 2x > 0, so (x+1)^2 > x^2 + 1. Quantity B is greater."],
    ["GRE Quant MCQ", "Easy", "MCQ", "If 3x + 7 = 22, what is the value of 6x + 4?", "", "15", "30", "34", "40", "48", "", "C", "3x = 15 => x = 5. 6(5) + 4 = 34."],
    ["GRE Quant MCQ", "Medium", "MCQ", "A circle has an area of 36π. What is its circumference?", "", "6π", "12π", "18π", "24π", "36π", "", "B", "Area = π r^2 = 36π => r = 6. Circumference = 2π r = 12π."],
    ["GRE Quant Numeric", "Easy", "NUMERIC_ENTRY", "A store discounted a $80 jacket by 25%. What is the sale price in dollars?", "", "", "", "", "", "", "", "60", "25% of $80 is $20. Sale price = 80 - 20 = $60."],
    ["GRE Quant Numeric", "Medium", "NUMERIC_ENTRY", "If the average (arithmetic mean) of 5, 10, 15, and x is 12, what is the value of x?", "", "", "", "", "", "", "", "19", "Sum = 5 + 10 + 15 + x = 30 + x. Total = 4 * 12 = 48. x = 48 - 30 = 19."],
    ["GRE Quant Multi", "Medium", "MULTIPLE_CHOICE_MULTI", "Which of the following integers are prime numbers?", "", "2", "9", "15", "17", "21", "", "A,D", "2 and 17 are prime numbers. 9 (3x3), 15 (3x5), and 21 (3x7) are composite."],
    ["GRE Quant Multi", "Hard", "MULTIPLE_CHOICE_MULTI", "If n is an even integer, which of the following MUST be even?", "", "n + 2", "3n", "n^2", "n + 1", "n / 2", "", "A,B,C", "n+2, 3n, and n^2 are always even when n is even."],
    ["GRE Quant QC", "Hard", "QUANTITATIVE_COMPARISON", "Triangle ABC has side lengths 5, 12, and 13.\nQuantity A: Area of triangle ABC\nQuantity B: 30", "", "Quantity A is greater.", "Quantity B is greater.", "The two quantities are equal.", "The relationship cannot be determined from the information given.", "", "", "C", "5^2 + 12^2 = 13^2, so it is a right triangle. Area = (1/2)*5*12 = 30."],
    ["GRE Quant MCQ", "Hard", "MCQ", "How many positive factors does the number 72 have?", "", "6", "8", "10", "12", "14", "", "D", "72 = 2^3 * 3^2. Number of factors = (3+1)*(2+1) = 4 * 3 = 12."]
]
create_excel("sample_quant_questions.xlsx", "Quant Questions", quant_rows)

# 3. Verbal Questions Excel (10 rows)
verbal_rows = [
    ["GRE Sentence Equivalence", "Medium", "SENTENCE_EQUIVALENCE", "Despite the CEO's ____ attitude during negotiations, the union delegates managed to secure a favorable contract.", "", "belligerent", "conciliatory", "pugnacious", "pliable", "amicable", "deferential", "A,C", "Belligerent and pugnacious both mean hostile/aggressive, creating equivalent sentences."],
    ["GRE Sentence Equivalence", "Easy", "SENTENCE_EQUIVALENCE", "The novel's plot was so ____ that readers found it almost impossible to follow.", "", "convoluted", "straightforward", "intricate", "lucid", "candid", "pellucid", "A,C", "Convoluted and intricate both mean complex/difficult to follow."],
    ["GRE Text Completion 1-Blank", "Easy", "TEXT_COMPLETION", "The scientist was praised for her ____ demeanor, which put her anxious colleagues at ease.", "", "tranquil", "boisterous", "irascible", "", "", "", "A", "Tranquil means calm, which puts anxious colleagues at ease."],
    ["GRE Text Completion 1-Blank", "Medium", "TEXT_COMPLETION", "The proposal was met with ____ hostility from committee members who favored traditional methods.", "", "unmitigated", "fleeting", "dubious", "superficial", "lukewarm", "", "A", "Unmitigated hostility means absolute/unqualified hostility."],
    ["GRE Text Completion 2-Blank", "Medium", "TEXT_COMPLETION", "While the author's early works were characterized by a (i) ____ prose style, her later writings demonstrated a surprising (ii) ____.", "", "prolix", "succinct", "pedantic", "brevity", "verbosity", "obscurity", "A,D", "Early prolix (verbose) contrasts with later brevity (conciseness)."],
    ["GRE Text Completion 2-Blank", "Hard", "TEXT_COMPLETION", "The politician's speech was surprisingly (i) ____, avoiding the usual vague promises and offering (ii) ____ solutions.", "", "candid", "evasive", "recondite", "concrete", "abstract", "tenuous", "A,D", "Candid (honest) speech offers concrete (tangible) solutions."],
    ["GRE Text Completion 3-Blank", "Hard", "TEXT_COMPLETION", "The biographer sought to (i) ____ the legend of the composer, demonstrating that his famous modesty was a (ii) ____ designed to (iii) ____ his intense ambition.", "A", "debunk", "immortalize", "vindicate", "facade", "paragon", "catalyst", "mask", "accentuate", "deprecate", "A,D,G", "Debunk legend, modesty was a facade to mask ambition."],
    ["GRE Reading Comprehension", "Medium", "MCQ", "According to the passage, what is the main reason for the decline in the whale population during the 19th century?", "For centuries, marine mammals faced minimal threats until commercial whaling escalated in the 19th century. Advances in harpoon technology and steam-powered vessels enabled fleets to hunt whales far more efficiently than in prior eras.", "Climate change", "Commercial whaling and technology advances", "Loss of ocean prey", "Pollution from coastal factories", "Natural disease outbreaks", "", "B", "Passage highlights commercial whaling and technological advances in harpoons."],
    ["GRE Reading Comprehension", "Hard", "MCQ", "The author's tone toward the early economic policies can best be described as:", "Early economic policies favored industrial monopolies under the guise of national growth. However, contemporary records reveal that these policies disproportionately harmed small agricultural communities, stifling regional innovation while enriching a select few.", "Enthusiastic and supportive", "Critical and objective", "Indifferent and detached", "Apologetic and defensive", "Sarcastic and hostile", "", "B", "The author objectively criticizes the economic policies for harming communities."],
    ["GRE Sentence Equivalence", "Hard", "SENTENCE_EQUIVALENCE", "The critic's review was notoriously ____, leaving no aspect of the performance unspared from harsh rebuke.", "", "scathing", "laudatory", "vitriolic", "tame", "complimentary", "temperate", "A,C", "Scathing and vitriolic both mean severely critical."]
]
create_excel("sample_verbal_questions.xlsx", "Verbal Questions", verbal_rows)

# 4. Image Questions Excel (Question Images & Answer Option Images Edge Cases - 10 rows)
image_rows = [
    ["GRE Geometry", "Medium", "MCQ", "In the geometric diagram below, what is the area of the shaded region?", "", "12", "16", "24", "32", "48", "", "C", "Area = (1/2)*base*height = (1/2)*6*8 = 24.", "triangle_geometry_fig1.png", ""],
    ["GRE Coordinate Geometry", "Hard", "QUANTITATIVE_COMPARISON", "Refer to the coordinate plane figure.\nQuantity A: Slope of line L1\nQuantity B: Slope of line L2", "", "Quantity A is greater.", "Quantity B is greater.", "The two quantities are equal.", "The relationship cannot be determined from the information given.", "", "", "A", "Line L1 has slope 2 while line L2 has slope 0.5. 2 > 0.5.", "coordinate_grid_lines.png", ""],
    ["GRE Data Interpretation", "Medium", "MULTIPLE_CHOICE_MULTI", "Based on the bar graph figure, in which years did total revenue exceed $50 million?", "", "2018", "2019", "2020", "2021", "2022", "", "B,D", "Bar heights for 2019 ($55M) and 2021 ($62M) exceed $50M.", "bar_chart_revenue.png", ""],
    ["GRE Quant Image Options", "Hard", "MCQ", "Which of the option diagrams represents the correct graph of y = |x - 3|?", "", "Graph Option A", "Graph Option B", "Graph Option C", "Graph Option D", "", "", "B", "y = |x - 3| has a V-shape vertex at (3,0).", "question_graph_v.png", "opt_graph_a.png, opt_graph_b.png, opt_graph_c.png, opt_graph_d.png"],
    ["GRE Geometry Circle", "Easy", "MCQ", "In circle O shown in the figure, what is the measure of angle AOB?", "", "30°", "45°", "60°", "90°", "120°", "", "C", "Angle AOB subtends an arc of 60°.", "circle_angle_fig.png", ""],
    ["GRE Multi-Figure", "Hard", "QUANTITATIVE_COMPARISON", "Refer to Figure 1 and Figure 2.\nQuantity A: Perimeter of Figure 1\nQuantity B: Perimeter of Figure 2", "", "Quantity A is greater.", "Quantity B is greater.", "The two quantities are equal.", "The relationship cannot be determined from the information given.", "", "", "C", "Both figures have equal perimeter of 24 units.", "fig1_square.png, fig2_rectangle.png", ""],
    ["GRE Image Choices", "Medium", "MCQ", "Which of the vector diagrams correctly shows the resultant vector R = A + B?", "", "Vector Choice A", "Vector Choice B", "Vector Choice C", "Vector Choice D", "", "", "A", "Vector A tip-to-tail with Vector B yields resultant R in choice A.", "", "vector_a.png, vector_b.png, vector_c.png, vector_d.png"],
    ["GRE Quant Figure Numeric", "Medium", "NUMERIC_ENTRY", "In the right triangle shown in the figure, what is the length of side x?", "", "", "", "", "", "", "", "15", "x = sqrt(9^2 + 12^2) = sqrt(81 + 144) = sqrt(225) = 15.", "right_triangle_xyz.png", ""],
    ["GRE Verbal Image Comprehension", "Medium", "MCQ", "Which architectural component highlighted in the diagram corresponds to the 'buttress'?", "", "Component A", "Component B", "Component C", "Component D", "", "", "C", "Component C shows the flying buttress support structure.", "gothic_cathedral_diagram.png", ""],
    ["GRE Data Chart", "Easy", "MCQ", "Based on the pie chart, what percentage of total expenses was spent on research?", "", "15%", "25%", "35%", "40%", "50%", "", "B", "The pie chart sector for Research indicates 25%.", "expense_pie_chart.png", ""]
]
create_excel("sample_image_questions.xlsx", "Image Questions", image_rows)

print("Successfully generated 4 Excel templates in /home/ramji/Desktop/GRE/gre/excels:")
print(" 1. sample_awa_questions.xlsx")
print(" 2. sample_quant_questions.xlsx")
print(" 3. sample_verbal_questions.xlsx")
print(" 4. sample_image_questions.xlsx")
